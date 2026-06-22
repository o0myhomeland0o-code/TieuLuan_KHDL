import pandas as pd
import matplotlib.pyplot as plt

from sklearn.cluster import KMeans
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, classification_report
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# 1. ĐỌC DỮ LIỆU

FILE_INPUT = "du_lieu_diem.xlsx"
df = pd.read_excel(FILE_INPUT)

# 2. TÍNH GPA HỆ 10

cac_mon = [
    'Toan',
    'NguVan',
    'NgoaiNgu',
    'VatLy',
    'HoaHoc',
    'SinhHoc',
    'LichSu',
    'DiaLy'
]

df['GPA'] = df[cac_mon].mean(axis=1).round(2)

# 3. PHÂN LOẠI HỌC LỰC

def phan_loai(gpa):
    if gpa >= 8:
        return "Giỏi"
    elif gpa >= 6.5:
        return "Khá"
    else:
        return "Trung bình"

df['Xep_loai'] = df['GPA'].apply(phan_loai)

# 4. ĐÁNH GIÁ

def danh_gia(gpa):
    if gpa >= 8:
        return "Học lực xuất sắc"
    elif gpa >= 6.5:
        return "Học lực khá"
    else:
        return "Cần cải thiện thêm"

df['Danh_gia'] = df['GPA'].apply(danh_gia)

# 5. SẮP XẾP DỮ LIỆU

df_sorted = df.sort_values(
    by='GPA',
    ascending=False
).reset_index(drop=True)

df_gioi = df_sorted[df_sorted['Xep_loai'] == 'Giỏi']
df_kha = df_sorted[df_sorted['Xep_loai'] == 'Khá']
df_tb = df_sorted[df_sorted['Xep_loai'] == 'Trung bình']

total_count = len(df_sorted)

# 6. BIỂU ĐỒ THỐNG KÊ HỌC LỰC

counts = [
    len(df_gioi),
    len(df_kha),
    len(df_tb)
]

labels = [
    'Giỏi',
    'Khá',
    'Trung bình'
]

colors = [
    '#70AD47',
    '#FFC000',
    '#ED7D31'
]

fig, ax = plt.subplots(figsize=(7, 5))

bars = ax.bar(
    labels,
    counts,
    color=colors
)

ax.set_title("Thống kê học lực học sinh")
ax.set_ylabel("Số lượng học sinh")

for bar in bars:
    y = bar.get_height()
    ax.text(
        bar.get_x() + bar.get_width()/2,
        y + 1,
        str(int(y)),
        ha='center'
    )

plt.tight_layout()

chart_path = "thong_ke.png"

plt.savefig(chart_path, dpi=150)
plt.close()

# 7. PHÂN CỤM K-MEANS

X_cluster = df[
    ['GPA',
     'SoGioHocMoiTuan',
     'SoBuoiNghiHoc']
]

kmeans = KMeans(
    n_clusters=3,
    random_state=42,
    n_init=10
)

df['Cluster'] = kmeans.fit_predict(X_cluster)

# RANDOM FOREST DỰ ĐOÁN ĐỖ / TRƯỢT

encoder = LabelEncoder()

df['KetQua_ML'] = encoder.fit_transform(
    df['KetQua']
)

features = [
    'Toan',
    'NguVan',
    'NgoaiNgu',
    'VatLy',
    'HoaHoc',
    'SinhHoc',
    'LichSu',
    'DiaLy',
    'SoGioHocMoiTuan',
    'SoBuoiNghiHoc'
]

X = df[features]
y = df['KetQua_ML']

X_train, X_test, y_train, y_test = train_test_split(
    X,
    y,
    test_size=0.2,
    random_state=42
)

# RANDOM FOREST

rf_model = RandomForestClassifier(
    n_estimators=100,
    random_state=42
)

rf_model.fit(X_train, y_train)

y_pred_rf = rf_model.predict(X_test)

accuracy_rf = accuracy_score(
    y_test,
    y_pred_rf
)

print("=" * 50)
print("Độ chính xác Random Forest:", round(accuracy_rf, 4))
print("=" * 50)

# LOGISTIC REGRESSION

log_model = LogisticRegression(
    max_iter=1000
)

log_model.fit(X_train, y_train)

y_pred_log = log_model.predict(X_test)

accuracy_log = accuracy_score(
    y_test,
    y_pred_log
)

print("Độ chính xác Logistic Regression:",
      round(accuracy_log, 4))

# BIỂU ĐỒ PHÂN BỐ GPA

plt.figure(figsize=(8, 5))

plt.hist(
    df['GPA'],
    bins=10,
    edgecolor='black'
)

plt.xlabel("Điểm trung bình")
plt.ylabel("Số học sinh")
plt.title("Biểu đồ phân bố điểm trung bình")

plt.grid(True)

plt.tight_layout()

plt.savefig("phan_bo_gpa.png")

plt.close()

# TOP 10 HỌC SINH GPA CAO NHẤT

top10 = df_sorted.head(10)

plt.figure(figsize=(10, 5))

plt.bar(
    top10['MaHS'],
    top10['GPA']
)

plt.xticks(rotation=45)

plt.xlabel("Mã học sinh")
plt.ylabel("Điểm trung bình")

plt.title("Top 10 học sinh có điểm trung bình cao nhất")

plt.tight_layout()

plt.savefig("top10_gpa.png")

plt.close()

# BIỂU ĐỒ K-MEANS

plt.figure(figsize=(7, 5))

plt.scatter(
    df['GPA'],
    df['SoGioHocMoiTuan'],
    c=df['Cluster']
)

plt.xlabel("Điểm trung bình")
plt.ylabel("Số giờ học mỗi tuần")

plt.title("Kết quả phân cụm K-Means")

plt.tight_layout()

plt.savefig("kmeans_cluster.png")

plt.close()

# IN TOP 10 HỌC SINH

top10_hs = df_sorted.head(10)

print("\nTOP 10 HỌC SINH CÓ GPA CAO NHẤT")

print(
    top10_hs[
        ['MaHS', 'GPA', 'Xep_loai']
    ]
)


# 8. THIẾT LẬP EXCEL

wb = Workbook()

COLOR_MAP = {
    'Giỏi': 'E2EFDA',
    'Khá': 'FFF2CC',
    'Trung bình': 'FCE4D6'
}

HEADER_FILL = PatternFill(
    'solid',
    start_color='2F5597',
    end_color='2F5597'
)

HEADER_FONT = Font(
    name='Segoe UI',
    size=11,
    bold=True,
    color='FFFFFF'
)

TEXT_FONT = Font(
    name='Segoe UI',
    size=10
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# SHEET 1 - XẾP HẠNG HỌC SINH

ws1 = wb.active
ws1.title = "Xep_hang_hoc_sinh"

headers = [
    "STT",
    "MaHS",
    "GioiTinh",
    "GPA He 10",
    "Xep loai",
    "Danh gia"
]

for i, text in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=i, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = thin_border

for idx, row in df_sorted.iterrows():

    fill = PatternFill(
        'solid',
        start_color=COLOR_MAP[row['Xep_loai']],
        end_color=COLOR_MAP[row['Xep_loai']]
    )

    values = [
        idx + 1,
        row['MaHS'],
        row['GioiTinh'],
        row['GPA'],
        row['Xep_loai'],
        row['Danh_gia']
    ]

    for c, val in enumerate(values, 1):

        cell = ws1.cell(
            row=idx + 2,
            column=c,
            value=val
        )

        cell.fill = fill
        cell.border = thin_border
        cell.font = TEXT_FONT

# SHEET 2 - THỐNG KÊ

ws2 = wb.create_sheet("Thong_ke")

headers2 = [
    "Nhom",
    "So luong",
    "Ty le"
]

for i, text in enumerate(headers2, 1):

    cell = ws2.cell(row=1, column=i, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = thin_border

tk_data = [
    ("Giỏi", len(df_gioi)),
    ("Khá", len(df_kha)),
    ("Trung bình", len(df_tb))
]

for idx, (nhom, count) in enumerate(tk_data):

    ws2.cell(idx + 2, 1, nhom)
    ws2.cell(idx + 2, 2, count)
    ws2.cell(idx + 2, 3, count / total_count)

img = OpenpyxlImage(chart_path)
ws2.add_image(img, "E2")

# SHEET 3 - KMEANS

ws3 = wb.create_sheet("Phan_cum_KMeans")

headers3 = [
    "MaHS",
    "GPA",
    "SoGioHocMoiTuan",
    "SoBuoiNghiHoc",
    "Cluster"
]

for i, text in enumerate(headers3, 1):

    cell = ws3.cell(row=1, column=i, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = thin_border

for idx, row in df.iterrows():

    values = [
        row['MaHS'],
        row['GPA'],
        row['SoGioHocMoiTuan'],
        row['SoBuoiNghiHoc'],
        row['Cluster']
    ]

    for c, val in enumerate(values, 1):

        cell = ws3.cell(
            row=idx + 2,
            column=c,
            value=val
        )

        cell.border = thin_border

# SHEET 4 - KẾT QUẢ ĐỖ / TRƯỢT

ws4 = wb.create_sheet("Danh_sach_Do_Truot")

headers4 = [
    "MaHS",
    "KhoiXetTuyen",
    "TruongDangKy",
    "DiemXetTuyen",
    "KetQua"
]

for i, text in enumerate(headers4, 1):

    cell = ws4.cell(row=1, column=i, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = thin_border

for idx, row in df.iterrows():

    values = [
        row['MaHS'],
        row['KhoiXetTuyen'],
        row['TruongDangKy'],
        row['DiemXetTuyen'],
        row['KetQua']
    ]

    for c, val in enumerate(values, 1):

        cell = ws4.cell(
            row=idx + 2,
            column=c,
            value=val
        )

        cell.border = thin_border

# SHEET 5 - RANDOM FOREST

ws5 = wb.create_sheet("Random_Forest")

ws5['A1'] = "Mô hình"
ws5['B1'] = "Độ chính xác"

ws5['A2'] = "Random Forest"
ws5['B2'] = round(accuracy_rf, 4)

ws5['A3'] = "Logistic Regression"
ws5['B3'] = round(accuracy_log, 4)

for row in range(1, 4):
    for col in range(1, 3):

        cell = ws5.cell(row=row, column=col)

        if row == 1:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        cell.border = thin_border

# SHEET 6 - BẢNG DỮ LIỆU GỐC

ws6 = wb.create_sheet("Bang_du_lieu")

for col_num, column_name in enumerate(df.columns, 1):

    cell = ws6.cell(
        row=1,
        column=col_num,
        value=column_name
    )

    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = thin_border

for row_num, row in enumerate(df.values, 2):

    for col_num, value in enumerate(row, 1):

        cell = ws6.cell(
            row=row_num,
            column=col_num,
            value=value
        )

        cell.border = thin_border



# CHÈN CÁC BIỂU ĐỒ

img1 = OpenpyxlImage("phan_bo_gpa.png")
ws2.add_image(img1, "E20")

img2 = OpenpyxlImage("top10_gpa.png")
ws2.add_image(img2, "E40")

img3 = OpenpyxlImage("kmeans_cluster.png")
ws3.add_image(img3, "H2")



# TỰ ĐỘNG CHỈNH ĐỘ RỘNG CỘT

for ws in wb.worksheets:

    for col in ws.columns:

        max_length = 0
        column = col[0].column_letter

        for cell in col:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except:
                pass

        ws.column_dimensions[column].width = max_length + 4

# LƯU FILE EXCEL

OUTPUT_PATH = "danh_gia.xlsx"

wb.save(OUTPUT_PATH)

print("=" * 60)
print("HOÀN THÀNH PHÂN TÍCH DỮ LIỆU")
print("Đã tạo file:", OUTPUT_PATH)
print("=" * 60)